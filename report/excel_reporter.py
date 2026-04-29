"""Excel 测试报告生成器

使用 openpyxl 生成测试报告 Excel 文件。
"""

from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("需要安装 openpyxl 库: pip install openpyxl")


class ExcelReporter:
    """Excel 测试报告生成器"""

    # 样式定义
    HEADER_FONT = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

    PASS_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    PASS_FONT = Font(color='006100')

    FAIL_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    FAIL_FONT = Font(color='9C0006')

    WARN_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    WARN_FONT = Font(color='9C6500')

    NORMAL_FONT = Font(name='微软雅黑', size=10)
    TITLE_FONT = Font(name='微软雅黑', bold=True, size=14)
    SUBTITLE_FONT = Font(name='微软雅黑', bold=True, size=11)

    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def export(self, file_path: str, test_items: List[Dict[str, Any]],
               test_log: str = "", test_start_time: Optional[datetime] = None,
               test_end_time: Optional[datetime] = None) -> str:
        """
        导出测试报告到 Excel 文件

        参数:
            file_path: 保存路径
            test_items: 测试项列表
            test_log: 测试日志
            test_start_time: 测试开始时间
            test_end_time: 测试结束时间

        返回:
            保存的文件路径
        """
        wb = Workbook()

        # Sheet 1: 测试概要
        self._create_summary_sheet(wb, test_items, test_start_time, test_end_time)

        # Sheet 2: 详细结果
        self._create_detail_sheet(wb, test_items)

        # Sheet 3: 测试日志
        self._create_log_sheet(wb, test_log)

        # 保存文件
        wb.save(file_path)
        return file_path

    def _create_summary_sheet(self, wb: Workbook, items: List[Dict],
                              start_time: Optional[datetime],
                              end_time: Optional[datetime]):
        """创建测试概要 Sheet"""
        ws = wb.active
        ws.title = "测试概要"

        # 标题
        ws.merge_cells('A1:F1')
        ws['A1'] = "协议一致性测试报告"
        ws['A1'].font = self.TITLE_FONT
        ws['A1'].alignment = Alignment(horizontal='center')

        # 统计数据
        total = len(items)
        passed = sum(1 for item in items if item.get('test_result') == '通过')
        failed = sum(1 for item in items if item.get('test_result') == '失败')
        timeout = sum(1 for item in items if item.get('test_result') == '超时')
        skipped = sum(1 for item in items if item.get('test_result') == '未测')

        pass_rate = (passed / total * 100) if total > 0 else 0

        # 概要信息
        summary_data = [
            ("测试信息", "", "", "", "", ""),
            ("测试开始时间", start_time.strftime("%Y-%m-%d %H:%M:%S") if start_time else "未记录", "", "", "", ""),
            ("测试结束时间", end_time.strftime("%Y-%m-%d %H:%M:%S") if end_time else "未记录", "", "", "", ""),
            ("", "", "", "", "", ""),
            ("测试统计", "", "", "", "", ""),
            ("总步骤数", total, "", "通过数", passed, ""),
            ("失败数", failed, "", "超时数", timeout, ""),
            ("未测数", skipped, "", "通过率", f"{pass_rate:.1f}%", ""),
        ]

        for row_idx, row_data in enumerate(summary_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.NORMAL_FONT
                cell.border = self.THIN_BORDER
                if col_idx == 1 or col_idx == 4:
                    cell.font = self.SUBTITLE_FONT
                    cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

        # 设置列宽
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15

    def _create_detail_sheet(self, wb: Workbook, items: List[Dict]):
        """创建详细结果 Sheet"""
        ws = wb.create_sheet("详细结果")

        # 表头
        headers = ["序号", "测试项名称", "发送帧", "匹配规则", "响应帧", "匹配结果", "状态"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # 数据行
        for row_idx, item in enumerate(items, start=2):
            # 序号
            ws.cell(row=row_idx, column=1, value=row_idx - 1).font = self.NORMAL_FONT

            # 名称
            ws.cell(row=row_idx, column=2, value=item.get('name', '')).font = self.NORMAL_FONT

            # 发送帧
            frame_hex = item.get('frame_hex', '')
            ws.cell(row=row_idx, column=3, value=self._format_hex(frame_hex)).font = self.NORMAL_FONT

            # 匹配规则
            match_rule = item.get('match_rule', '')
            ws.cell(row=row_idx, column=4, value=self._format_hex(match_rule)).font = self.NORMAL_FONT

            # 响应帧
            response = item.get('response_frame', '')
            ws.cell(row=row_idx, column=5, value=self._format_hex(response)).font = self.NORMAL_FONT

            # 匹配结果
            result = item.get('test_result', '未测')
            result_cell = ws.cell(row=row_idx, column=6, value=result)
            result_cell.font = self.NORMAL_FONT
            if result == '通过':
                result_cell.fill = self.PASS_FILL
                result_cell.font = self.PASS_FONT
            elif result == '失败':
                result_cell.fill = self.FAIL_FILL
                result_cell.font = self.FAIL_FONT
            elif result == '超时':
                result_cell.fill = self.WARN_FILL
                result_cell.font = self.WARN_FONT

            # 状态
            status = item.get('status', '待测')
            ws.cell(row=row_idx, column=7, value=status).font = self.NORMAL_FONT

            # 设置边框
            for col_idx in range(1, 8):
                ws.cell(row=row_idx, column=col_idx).border = self.THIN_BORDER

        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12

    def _create_log_sheet(self, wb: Workbook, log_text: str):
        """创建测试日志 Sheet"""
        ws = wb.create_sheet("测试日志")

        # 标题
        ws.merge_cells('A1:A1')
        ws['A1'] = "测试执行日志"
        ws['A1'].font = self.SUBTITLE_FONT

        # 日志内容
        if log_text:
            lines = log_text.split('\n')
            for row_idx, line in enumerate(lines, start=3):
                cell = ws.cell(row=row_idx, column=1, value=line)
                cell.font = Font(name='Consolas', size=9)

        # 设置列宽
        ws.column_dimensions['A'].width = 120

    def _format_hex(self, hex_str: str) -> str:
        """格式化 HEX 字符串（添加空格分隔）"""
        if not hex_str:
            return ""
        clean = hex_str.replace(" ", "").upper()
        return " ".join(clean[i:i+2] for i in range(0, len(clean), 2))

# -*- coding: utf-8 -*-
"""批量解析标签页"""
import io
import json
import re
from typing import List
from nicegui import ui
from web.components.parse_table import ParseTable
from web.protocol_registry import make_parser
from web.frame_extractor import extract_frames_for_protocol


class BatchParseTab:

    CSG_MONITOR_PREFIX = "-> 接收机 Has Get"
    CSG_MONITOR_HEADER_BYTES = 15

    def __init__(self, protocol_selector):
        self.protocol_selector = protocol_selector
        self.current_protocol = 0
        self.parser = make_parser(0)
        self._batch_input = None
        self._summary_table = None
        self._detail_table = None
        self._results = []
        self._parse_level = "auto"

    def build(self):
        # 注入 html2canvas 库（用于导出图片）
        ui.add_head_html(
            '<script src="https://cdn.jsdelivr.net/npm/html2canvas@1.4.1/dist/html2canvas.min.js"></script>'
        )
        # 全局阴影与圆角卡片样式
        ui.add_head_html('''
        <style>
            .batch-card { border-radius: 12px !important; box-shadow: 0 2px 12px rgba(0,0,0,0.08) !important; }
            .batch-card-elevated { border-radius: 12px !important; box-shadow: 0 4px 16px rgba(0,0,0,0.12) !important; }
            .batch-section-title { font-weight: 600; letter-spacing: 0.3px; }
            .batch-status-bar {
                display: flex; align-items: center; justify-content: space-between;
                padding: 6px 16px; border-top: 1px solid #e0e0e0;
                background: #f8f9fa; border-radius: 0 0 12px 12px;
                font-size: 12px; color: #666;
            }
            .batch-divider { border-top: 1px solid #e0e0e0; margin: 4px 0 8px 0; }
        </style>
        ''')

        with ui.column().classes("w-full").style("height: 100%; display: flex; flex-direction: column; overflow: hidden;"):
            # ── 输入区 ──
            with ui.card().classes("w-full batch-card").style("margin: 8px 12px;"):
                # 标题行 + 分割线
                with ui.row().classes("w-full items-center").style("gap: 6px; padding: 8px 12px 0 12px;"):
                    ui.icon("playlist_add", size="20px", color="primary")
                    ui.label("批量输入报文").classes("batch-section-title text-subtitle1")
                    ui.label("(支持监控日志/纯HEX)").classes("text-caption text-grey-6")
                ui.html('<div class="batch-divider"></div>').style("margin: 0 12px;")

                self._batch_input = ui.textarea(
                    placeholder="粘贴监控日志或十六进制报文，每行一帧",
                ).classes("w-full").props('outlined dense rows=6').style(
                    "margin: 0 12px; font-family: monospace; font-size: 12px; min-height: 100px;"
                    "border-radius: 8px;"
                )

                # 按钮栏
                with ui.row().classes("w-full items-center").style("padding: 8px 12px 12px 12px; gap: 8px;"):
                    ui.button("批量解析", icon="playlist_add", on_click=self._do_batch_parse) \
                        .props("color=primary unelevated").classes("text-weight-medium")
                    # 分割线
                    ui.separator().props("vertical").style("height: 24px; margin: 0 4px;")
                    ui.button("导出 Excel", icon="download", on_click=self._export_excel) \
                        .props("color=green-7 flat").classes("text-weight-medium")
                    ui.button("导出 JSON", icon="download", on_click=self._export_json) \
                        .props("color=blue-7 flat").classes("text-weight-medium")
                    ui.button("导出图片", icon="download", on_click=self._export_image) \
                        .props("color=orange-7 flat").classes("text-weight-medium")
                    ui.button("清空", icon="clear_all", on_click=self._clear) \
                        .props("flat color=grey-6")
                    ui.space()
                    ui.label("解析级别:").classes("text-caption text-grey-6")
                    ui.select(
                        options={"auto": "自动", "fc_pb": "FC+PB", "fc_only": "仅FC", "app": "应用层"},
                        value="auto",
                    ).props("outlined dense").style("width: 120px; border-radius: 8px;").bind_value(self, '_parse_level')

            # ── 结果区：左右分割 ──
            with ui.splitter(value=45).classes("w-full").style("flex: 1; margin: 0 12px 8px 12px;") as splitter:
                with splitter.before:
                    with ui.card().classes("w-full h-full batch-card-elevated").style("overflow: hidden;"):
                        # 摘要标题栏
                        with ui.row().classes("w-full items-center").style("padding: 8px 12px; gap: 6px; border-bottom: 1px solid #e8e8e8;"):
                            ui.icon("table_chart", size="18px", color="primary")
                            ui.label("批量解析摘要").classes("batch-section-title text-subtitle1")
                            ui.space()
                            self._frame_count_label = ui.label("共 0 帧").classes("text-caption text-grey-6")
                        columns = [
                            {"name": "idx", "label": "#", "field": "idx", "width": "40px", "align": "center"},
                            {"name": "status", "label": "状态", "field": "status", "width": "50px", "align": "center"},
                            {"name": "len", "label": "长度", "field": "len", "width": "50px", "align": "right"},
                            {"name": "proto", "label": "协议/类型", "field": "proto"},
                            {"name": "summary", "label": "摘要", "field": "summary"},
                        ]
                        self._summary_table = ui.table(
                            columns=columns,
                            rows=[],
                            row_key="id",
                            selection="single",
                        ).classes("w-full").props("flat bordered separator=cell dense")
                        self._summary_table.style("max-height: calc(100vh - 280px);")
                        self._summary_table.on('rowClick', self._on_row_click)

                with splitter.after:
                    with ui.card().classes("w-full h-full batch-card-elevated").style(
                        "overflow: hidden; display: flex; flex-direction: column;"
                    ):
                        # 详细解析标题栏
                        with ui.row().classes("w-full items-center").style("padding: 8px 12px; gap: 6px; border-bottom: 1px solid #e8e8e8;"):
                            ui.icon("zoom_in", size="18px", color="primary")
                            ui.label("选中帧详细解析").classes("batch-section-title text-subtitle1")
                            ui.space()
                            ui.button(icon="content_copy", on_click=self._copy_detail) \
                                .props("flat dense round size=sm").tooltip("复制全部").style("color: #666;")
                        self._detail_hex = ui.label("点击左侧摘要行查看详细解析").classes(
                            "text-caption text-grey-5"
                        ).style("padding: 8px 12px; font-family: monospace; word-break: break-all;")
                        self._detail_hex.set_visibility(False)
                        detail_columns = [
                            {"name": "field", "label": "字段", "field": "field", "align": "left"},
                            {"name": "raw", "label": "原始值", "field": "raw", "align": "left"},
                            {"name": "parsed", "label": "解析值", "field": "parsed", "align": "left"},
                            {"name": "comment", "label": "说明", "field": "comment", "align": "left"},
                        ]
                        self._detail_table = ui.table(
                            columns=detail_columns,
                            rows=[],
                        ).classes("w-full").props("flat bordered separator=cell dense virtual-scroll").style("flex: 1;")

            # ── 底部状态栏 ──
            with ui.card().classes("w-full").style("margin: 0 12px 8px 12px; border-radius: 8px; border: 1px solid #e0e0e0;"):
                with ui.row().classes("w-full items-center").style("padding: 6px 16px; gap: 16px;"):
                    ui.icon("info", size="16px", color="grey-6")
                    self._status_label = ui.label("就绪 — 等待输入报文").classes("text-caption text-grey-6")
                    ui.space()
                    self._status_stats = ui.label("").classes("text-caption text-grey-7")

    def on_protocol_change(self, protocol_idx: int):
        self.current_protocol = protocol_idx
        self.parser = make_parser(protocol_idx)
        self._results = []
        if self._summary_table:
            self._summary_table.rows = []
        self._clear_detail()

    def _do_batch_parse(self):
        text = self._batch_input.value or ""
        if not text.strip():
            ui.notify("请输入报文", type="warning")
            return

        if self.current_protocol == 9:
            text = self._strip_csg_monitor_prefix(text)

        frame_hexes = extract_frames_for_protocol(text, self.current_protocol)
        if not frame_hexes:
            ui.notify("未找到有效帧数据", type="warning")
            return

        self._results = []
        summary_rows = []

        for idx, frame_hex in enumerate(frame_hexes):
            try:
                clean = re.sub(r'[^0-9A-Fa-f]', '', frame_hex).upper()
                if len(clean) % 2 != 0:
                    clean = clean[:-1]
                frame_bytes = bytes.fromhex(clean)

                if self.current_protocol == 9:
                    result = self.parser.parse_to_table(
                        frame_bytes,
                        parse_level=getattr(self, '_parse_level', 'auto'),
                        pb_frame_type='sof',
                    )
                else:
                    result = self.parser.parse_to_table(frame_bytes)

                status = "成功" if result and result[0][0] != "❌ 解析失败" else "失败"
                proto_name = self._get_frame_summary(result)

                self._results.append({
                    "id": idx,
                    "frame_bytes": frame_bytes,
                    "result": result,
                    "status": status,
                })

                summary_rows.append({
                    "id": idx,
                    "idx": idx + 1,
                    "status": "✅" if status == "成功" else "❌",
                    "len": len(frame_bytes),
                    "proto": proto_name,
                    "summary": self._extract_summary(result)[:100],
                })
            except Exception as ex:
                self._results.append({
                    "id": idx,
                    "frame_bytes": b"",
                    "result": [],
                    "status": "错误",
                    "error": str(ex),
                })
                summary_rows.append({
                    "id": idx,
                    "idx": idx + 1,
                    "status": "❌",
                    "len": 0,
                    "proto": "解析错误",
                    "summary": str(ex)[:100],
                })

        self._summary_table.rows = summary_rows
        self._clear_detail()
        # 更新状态栏和帧计数
        success_count = sum(1 for r in self._results if r["status"] == "成功")
        fail_count = len(self._results) - success_count
        self._frame_count_label.set_text(f"共 {len(summary_rows)} 帧")
        self._status_label.set_text(
            f"解析完成 — 共 {len(summary_rows)} 帧"
            f"（✅ {success_count} 成功"
            + (f"，❌ {fail_count} 失败" if fail_count else "")
            + "）"
        )
        ui.notify(f"批量解析完成：提取到 {len(frame_hexes)} 帧", type="positive")

    def _on_row_click(self, e):
        """NiceGUI rowClick 事件处理"""
        idx = -1
        try:
            if e.args:
                if len(e.args) >= 2:
                    val = e.args[1]
                    if isinstance(val, dict):
                        idx = val.get("id", -1)
                    elif isinstance(val, int):
                        idx = val
                elif len(e.args) >= 1:
                    row = e.args[0]
                    if isinstance(row, dict):
                        idx = row.get("id", -1)
        except Exception:
            pass

        if idx < 0:
            try:
                selected = self._summary_table.selected
                if selected and len(selected) > 0:
                    idx = selected[0].get("id", -1)
            except Exception:
                pass


        if 0 <= idx < len(self._results):
            item = self._results[idx]
            result = item["result"]
            frame_bytes = item["frame_bytes"]

            # 显示原始帧 HEX
            if frame_bytes:
                hex_str = " ".join(f"{b:02X}" for b in frame_bytes)
                self._detail_hex.set_text(f"原始帧: {hex_str}")
                self._detail_hex.set_visibility(True)
            else:
                self._detail_hex.set_visibility(False)

            # 填充表格数据
            rows = []
            for row in result:
                if len(row) >= 4:
                    field = str(row[0]) if row[0] else ""
                    raw = str(row[1]) if row[1] is not None else ""
                    parsed = str(row[2]) if row[2] is not None else ""
                    comment = str(row[3]) if row[3] is not None else ""
                    is_child = row[6] if len(row) > 6 else False
                    if is_child:
                        field = "  └ " + field
                    rows.append({
                        "field": field,
                        "raw": raw,
                        "parsed": parsed,
                        "comment": comment,
                    })
            self._detail_table.rows = rows
            return

    def _clear_detail(self):
        self._detail_table.rows = []
        self._detail_hex.set_visibility(False)
        self._detail_hex.set_text("点击左侧摘要行查看详细解析")

    def _copy_detail(self):
        idx = self._get_selected_idx()
        if idx is None or idx < 0 or idx >= len(self._results):
            ui.notify("请先选中一行", type="warning")
            return
        item = self._results[idx]
        lines = []
        for row in item["result"]:
            if len(row) >= 4:
                lines.append(f"{row[0]}\t{row[1]}\t{row[2]}\t{row[3]}")
        ui.clipboard.write("\n".join(lines))
        ui.notify("已复制到剪贴板", type="positive")

    def _get_selected_idx(self):
        try:
            selected = self._summary_table.selected
            if selected and len(selected) > 0:
                return selected[0].get("id", -1)
        except Exception:
            pass
        return -1

    def _strip_csg_monitor_prefix(self, text: str) -> str:
        prefix = self.CSG_MONITOR_PREFIX
        prefix_len = len(prefix)
        hex_only_line_re = re.compile(r'^[0-9A-Fa-f\s,\-]*$')
        out_lines = []
        for line in text.splitlines():
            pos = line.find(prefix)
            if pos == -1:
                if hex_only_line_re.match(line):
                    out_lines.append(line)
                continue
            after = line[pos + prefix_len:]
            tokens = re.findall(r'[0-9A-Fa-f]{1,2}', after)
            payload_tokens = tokens[self.CSG_MONITOR_HEADER_BYTES:]
            out_lines.append(' '.join(payload_tokens))
        return '\n'.join(out_lines)

    def _get_frame_summary(self, result: List) -> str:
        if not result:
            return "空"
        for row in result:
            field = row[0]
            if "AFN" in field or "帧类型" in field or "业务标识" in field or "功能码" in field:
                return str(row[3]) or str(row[2]) or field
        return result[0][0]

    def _extract_summary(self, result: List) -> str:
        if not result:
            return ""
        parts = []
        for row in result[:5]:
            if row[3]:
                parts.append(row[3])
        return "; ".join(parts)

    def _clear(self):
        self._batch_input.value = ""
        self._results = []
        self._summary_table.rows = []
        self._clear_detail()
        self._frame_count_label.set_text("共 0 帧")
        self._status_label.set_text("就绪 — 等待输入报文")
        self._status_stats.set_text("")

    # ── 导出 ──

    def _export_json(self):
        if not self._results:
            ui.notify("无数据可导出", type="warning")
            return
        data = []
        for item in self._results:
            data.append({
                "frame": item["frame_bytes"].hex().upper(),
                "status": item["status"],
                "result": [
                    {"field": r[0], "raw": r[1], "parsed": r[2], "comment": r[3]}
                    for r in item["result"] if len(r) >= 4
                ],
            })
        json_str = json.dumps(data, ensure_ascii=False, indent=2)
        ui.download(json_str.encode(), "batch_parse_result.json")
        ui.notify("已导出 JSON", type="positive")
    async def _export_image(self):
        """用 html2canvas 截取结果区，导出为 PNG"""
        if not self._results:
            ui.notify("无数据可导出", type="warning")
            return
        js_code = """
        (async () => {
            const splitter = document.querySelector('.q-splitter');
            if (!splitter) { return null; }
            const canvas = await html2canvas(splitter, {
                backgroundColor: '#ffffff',
                scale: 2,
                useCORS: true,
                logging: false,
            });
            return canvas.toDataURL('image/png');
        })()
        """
        try:
            result = await ui.run_javascript(js_code, timeout=10)
        except Exception:
            result = None
        if not result or not result.startswith("data:image/png;base64,"):
            ui.notify("截图失败", type="negative")
            return
        import base64
        b64 = result.split(",", 1)[1]
        img_bytes = base64.b64decode(b64)
        ui.download(img_bytes, "batch_parse_result.png")
        ui.notify("已导出图片", type="positive")


    def _export_excel(self):
        if not self._results:
            ui.notify("无数据可导出", type="warning")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            ui.notify("需要 openpyxl 库，请运行: pip install openpyxl", type="negative")
            return

        wb = Workbook()

        # ── 通用样式 ──
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # ========== Sheet1: 汇总表 ==========
        ws_summary = wb.active
        ws_summary.title = "汇总表"

        summary_headers = ["帧序号", "状态", "摘要", "原始数据", "错误信息", "包含详细数据"]
        for col, h in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for idx, item in enumerate(self._results):
            row_num = idx + 2
            summary = self._extract_summary(item["result"])[:100]
            raw_hex = item["frame_bytes"].hex().upper() if item["frame_bytes"] else ""
            error_msg = item.get("error", "")
            has_detail = "是" if item["result"] else "否"

            ws_summary.cell(row=row_num, column=1, value=idx + 1).border = thin_border
            status_cell = ws_summary.cell(row=row_num, column=2, value=item["status"])
            status_cell.border = thin_border
            if item["status"] == "成功":
                status_cell.font = Font(color="008000")
            elif item["status"] in ("失败", "错误"):
                status_cell.font = Font(color="FF0000")
            ws_summary.cell(row=row_num, column=3, value=summary).border = thin_border
            ws_summary.cell(row=row_num, column=4, value=raw_hex).border = thin_border
            ws_summary.cell(row=row_num, column=5, value=error_msg).border = thin_border
            ws_summary.cell(row=row_num, column=6, value=has_detail).border = thin_border

        for i, w in enumerate([10, 10, 40, 50, 30, 12], 1):
            ws_summary.column_dimensions[get_column_letter(i)].width = w

        # ========== Sheet2: 详细解析 ==========
        ws_detail = wb.create_sheet("详细解析")
        frame_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        frame_header_font = Font(bold=True, size=11, color="FFFFFF")
        field_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        field_header_font = Font(bold=True, size=10)

        current_row = 1

        for idx, item in enumerate(self._results):
            frame_num = idx + 1
            status = item["status"]
            summary = self._extract_summary(item["result"])[:80]

            # 帧标题行（绿色）
            ws_detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            title = f"第 {frame_num} 帧  |  状态: {status}  |  {summary}"
            frame_cell = ws_detail.cell(row=current_row, column=1, value=title)
            frame_cell.font = frame_header_font
            frame_cell.fill = frame_header_fill
            for c in range(1, 6):
                ws_detail.cell(row=current_row, column=c).border = thin_border
                ws_detail.cell(row=current_row, column=c).fill = frame_header_fill
            current_row += 1

            # 原始报文行
            if item["frame_bytes"]:
                ws_detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                hex_str = " ".join(f"{b:02X}" for b in item["frame_bytes"])
                ws_detail.cell(row=current_row, column=1, value=f"原始报文: {hex_str}").border = thin_border
                current_row += 1

            # 字段表头（浅蓝）
            field_headers = ["字段", "原始值", "解析值", "说明", "字节偏移"]
            for col, h in enumerate(field_headers, 1):
                cell = ws_detail.cell(row=current_row, column=col, value=h)
                cell.font = field_header_font
                cell.fill = field_header_fill
                cell.border = thin_border
            current_row += 1

            # 字段数据
            if item["result"]:
                for row in item["result"]:
                    field = str(row[0]) if len(row) > 0 else ""
                    raw = str(row[1]) if len(row) > 1 else ""
                    parsed = str(row[2]) if len(row) > 2 else ""
                    comment = str(row[3]) if len(row) > 3 else ""
                    byte_offset = ""
                    if len(row) > 5 and row[4] is not None and row[5] is not None:
                        byte_offset = f"{row[4]}-{row[5]}"

                    values = [field, raw, parsed, comment, byte_offset]
                    for col, val in enumerate(values, 1):
                        cell = ws_detail.cell(row=current_row, column=col, value=val)
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="top", wrap_text=(col == 4))
                    current_row += 1
            else:
                ws_detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                ws_detail.cell(row=current_row, column=1, value=f"错误: {item.get('error', '无详细解析数据')}").border = thin_border
                current_row += 1

            # 帧间分隔
            current_row += 1

        for i, w in enumerate([30, 20, 30, 40, 12], 1):
            ws_detail.column_dimensions[get_column_letter(i)].width = w

        # 保存到内存并下载
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        ui.download(buf.getvalue(), "batch_parse_result.xlsx")
        ui.notify(f"已导出 Excel: 汇总表 + 详细解析", type="positive")

"""
增强版批量解析导出工具

提供多格式导出（JSON、CSV）、智能文件名生成、验证和错误处理。
Sheet1: 汇总表（帧序号、状态、摘要、原始数据）
Sheet2: 每帧详细解析结果（按帧序号依次排列）
"""

import json
import csv
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional


class EnhancedBatchResultExporter:
    """
    增强版批量解析结果导出器。
    
    功能:
    - JSON 格式导出（完整数据 + 元数据）
    - CSV 格式导出（Sheet1 汇总 + Sheet2 详细解析，需安装 pandas + openpyxl）
    - 智能文件名（协议名称 + 时间戳）
    - 验证和错误处理
    """
    
    def __init__(self, export_dir: str = "."):
        """初始化导出器。"""
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(exist_ok=True)
    
    def export_to_json(self, batch_results: List[Dict[str, Any]],
                      protocol_name: str = "unknown",
                      output_file: Optional[str] = None) -> str:
        """
        导出到 JSON 格式。
        
        Args:
            batch_results: 批量解析结果列表
            protocol_name: 协议名称
            
        Returns:
            保存的 JSON 文件路径
        """
        timestamp = datetime.now()
        if output_file:
            file_path = Path(output_file)
            file_path.parent.mkdir(parents=True, exist_ok=True)
        else:
            timestamp_str = timestamp.strftime("%Y%m%d_%H%M%S")
            filename = f"batch_parse_{protocol_name}_{timestamp_str}.json"
            file_path = self.export_dir / filename
        
        export_data = {
            "metadata": {
                "export_time": timestamp.isoformat(),
                "protocol": protocol_name,
                "total_frames": len(batch_results),
                "version": "enhanced_batch_export_v1.0"
            },
            "results": []
        }
        
        for i, result in enumerate(batch_results):
            export_result = {
                "帧编号": i + 1,
                "状态": result.get("_status", result.get("状态", "")),
                "摘要": result.get("摘要", result.get("summary", "")),
                "原始数据": result.get("_input", result.get("原始数据", "")),
                "包含表格数据": "是" if "_table_data" in result else "否"
            }
            
            if "_table_data" not in result and "错误" in result:
                export_result["错误信息"] = result["错误"]
                
            export_data["results"].append(export_result)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)
            
        return str(file_path)
    
    def export_to_csv(self, batch_results: List[Dict[str, Any]],
                     protocol_name: str = "unknown") -> str:
        """
        导出到 CSV 格式（仅汇总表）。
        
        Args:
            batch_results: 批量解析结果列表
            protocol_name: 协议名称
            
        Returns:
            保存的 CSV 文件路径
            
        Raises:
            ImportError: 如果 pandas 未安装
        """
        try:
            import pandas as pd
        except ImportError:
            raise ImportError(
                "安装 pandas 后才能导出 CSV 格式。\n"
                "运行: pip install pandas"
            )
        
        timestamp = datetime.now()
        timestamp_str = timestamp.strftime("%Y%m%d_%H%M%S")
        filename = f"batch_parse_{protocol_name}_{timestamp_str}.csv"
        file_path = self.export_dir / filename
        
        csv_data = []
        for i, result in enumerate(batch_results):
            row = {
                "帧索引": i + 1,
                "状态": result.get("_status", result.get("状态", "")),
                "摘要": result.get("摘要", result.get("summary", "")),
                "原始数据": result.get("_input", result.get("原始数据", "")),
                "错误信息": result.get("错误", ""),
                "包含表格数据": "是" if "_table_data" in result else "否"
            }
            csv_data.append(row)
        
        df = pd.DataFrame(csv_data)
        df.to_csv(file_path, index=False, encoding='utf-8')
        
        return str(file_path)
    
    def export_to_excel(self, batch_results: List[Dict[str, Any]],
                       protocol_name: str = "unknown",
                       output_file: Optional[str] = None) -> str:
        """
        导出到 Excel 格式（Sheet1 汇总 + Sheet2 每帧详细解析）。
        
        Sheet1: 汇总表（帧序号、状态、摘要、原始数据、错误信息）
        Sheet2: 详细解析（按帧序号依次排列，每帧的所有解析字段）
        
        Args:
            batch_results: 批量解析结果列表
            protocol_name: 协议名称
            
        Returns:
            保存的 Excel 文件路径
            
        Raises:
            ImportError: 如果 pandas 或 openpyxl 未安装
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "安装 pandas 和 openpyxl 后才能导出 Excel 格式。\n"
                "运行: pip install pandas openpyxl"
            )
        
        if output_file:
            file_path = Path(output_file)
            file_path.parent.mkdir(parents=True, exist_ok=True)
        else:
            timestamp = datetime.now()
            timestamp_str = timestamp.strftime("%Y%m%d_%H%M%S")
            filename = f"batch_parse_{protocol_name}_{timestamp_str}.xlsx"
            file_path = self.export_dir / filename
        
        # 创建工作簿
        wb = Workbook()
        
        # ========== Sheet1: 汇总表 ==========
        ws_summary = wb.active
        ws_summary.title = "汇总表"
        
        # 汇总表样式
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 汇总表表头
        summary_headers = ["帧序号", "状态", "摘要", "原始数据", "错误信息", "包含详细数据"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # 汇总表数据
        for i, result in enumerate(batch_results):
            row_idx = i + 2
            
            ws_summary.cell(row=row_idx, column=1, value=i + 1).border = thin_border
            
            status = result.get("_status", result.get("状态", ""))
            status_cell = ws_summary.cell(row=row_idx, column=2, value=status)
            status_cell.border = thin_border
            if status == "成功":
                status_cell.font = Font(color="008000")
            elif status in ("失败", "异常"):
                status_cell.font = Font(color="FF0000")
            
            ws_summary.cell(row=row_idx, column=3, value=result.get("摘要", result.get("summary", ""))).border = thin_border
            ws_summary.cell(row=row_idx, column=4, value=result.get("_input", result.get("原始数据", ""))).border = thin_border
            ws_summary.cell(row=row_idx, column=5, value=result.get("错误", "")).border = thin_border
            ws_summary.cell(row=row_idx, column=6, value="是" if "_table_data" in result else "否").border = thin_border
        
        # 调整汇总表列宽
        summary_col_widths = [10, 10, 40, 50, 30, 12]
        for i, width in enumerate(summary_col_widths, 1):
            ws_summary.column_dimensions[get_column_letter(i)].width = width
        
        # ========== Sheet2: 详细解析 ==========
        ws_detail = wb.create_sheet("详细解析")
        
        # 详细解析样式
        frame_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        frame_header_font = Font(bold=True, size=11, color="FFFFFF")
        field_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        field_header_font = Font(bold=True, size=10)
        
        current_row = 1
        
        for i, result in enumerate(batch_results):
            frame_num = i + 1
            table_data = result.get("_table_data", [])
            
            # 帧标题行
            ws_detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            frame_title = f"第 {frame_num} 帧"
            status = result.get("_status", result.get("状态", ""))
            summary = result.get("摘要", result.get("summary", ""))
            frame_title += f"  |  状态: {status}  |  {summary}"
            
            frame_cell = ws_detail.cell(row=current_row, column=1, value=frame_title)
            frame_cell.font = frame_header_font
            frame_cell.fill = frame_header_fill
            frame_cell.alignment = Alignment(horizontal='left')
            for col in range(1, 6):
                ws_detail.cell(row=current_row, column=col).border = thin_border
                ws_detail.cell(row=current_row, column=col).fill = frame_header_fill
            current_row += 1
            
            # 原始数据行
            raw_data = result.get("_input", result.get("原始数据", ""))
            if raw_data:
                ws_detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                ws_detail.cell(row=current_row, column=1, value=f"原始报文: {raw_data}").border = thin_border
                current_row += 1
            
            # 字段表头
            field_headers = ["字段", "原始值", "解析值", "说明", "字节偏移"]
            for col, header in enumerate(field_headers, 1):
                cell = ws_detail.cell(row=current_row, column=col, value=header)
                cell.font = field_header_font
                cell.fill = field_header_fill
                cell.border = thin_border
            current_row += 1
            
            # 字段数据
            if table_data:
                for item in table_data:
                    # item 格式: (field_name, raw_value, parsed_value, comment, byte_start, byte_end)
                    field_name = str(item[0]) if len(item) > 0 else ""
                    raw_value = str(item[1]) if len(item) > 1 else ""
                    parsed_value = str(item[2]) if len(item) > 2 else ""
                    comment = str(item[3]) if len(item) > 3 else ""
                    
                    # 字节偏移
                    byte_offset = ""
                    if len(item) > 4 and len(item) > 5:
                        byte_start = item[4]
                        byte_end = item[5]
                        if byte_start is not None and byte_end is not None:
                            byte_offset = f"{byte_start}-{byte_end}"
                    
                    ws_detail.cell(row=current_row, column=1, value=field_name).border = thin_border
                    ws_detail.cell(row=current_row, column=2, value=raw_value).border = thin_border
                    ws_detail.cell(row=current_row, column=3, value=parsed_value).border = thin_border
                    ws_detail.cell(row=current_row, column=4, value=comment).border = thin_border
                    ws_detail.cell(row=current_row, column=5, value=byte_offset).border = thin_border
                    current_row += 1
            else:
                # 无详细数据时显示错误信息
                error_msg = result.get("错误", "无详细解析数据")
                ws_detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                ws_detail.cell(row=current_row, column=1, value=f"错误: {error_msg}").border = thin_border
                current_row += 1
            
            # 帧间分隔行
            current_row += 1
        
        # 调整详细解析列宽
        detail_col_widths = [30, 20, 30, 40, 12]
        for i, width in enumerate(detail_col_widths, 1):
            ws_detail.column_dimensions[get_column_letter(i)].width = width
        
        # 保存文件
        wb.save(file_path)
        
        return str(file_path)
    
    def validate_export_data(self, batch_results: List[Dict[str, Any]]) -> List[str]:
        """验证导出数据。"""
        errors = []
        
        if not batch_results:
            errors.append("批量解析结果为空")
            return errors
        
        for i, result in enumerate(batch_results):
            if not isinstance(result, dict):
                errors.append(f"第 {i+1} 帧结果不是字典类型")
            elif not result:
                errors.append(f"第 {i+1} 帧结果为空")
        
        return errors


# 示例用法
if __name__ == "__main__":
    print("=== 增强版批量解析导出工具测试 ===\n")
    
    sample_results = [
        {
            "_input": "68 0E 00 00 00 00 01 00 01 E8 00 05 EF 16",
            "_status": "成功",
            "摘要": "主站复位请求",
            "_table_data": [
                ["  协议控制信息", "68", "帧起始符", "固定值0x68", 0, 0],
                ["  长度域", "0E", "14字节", "用户数据长度", 1, 1],
                ["  控制域", "00 00 00 00", "主站→模块", "下行帧", 2, 5],
                ["  地址域", "01 00 01 E8", "地址: 0x010001E8", "集中器地址", 6, 9],
                ["  功能码", "00 05", "复位命令", "主站复位", 10, 11],
                ["  校验码", "EF", "校验和", "算术和校验", 12, 12],
                ["  结束符", "16", "帧结束符", "固定值0x16", 13, 13],
            ]
        },
        {
            "_input": "68 0E 00 00 00 00 01 02 01 28 00 03 01 20 00",
            "_status": "失败",
            "摘要": "查询数据失败",
            "错误": "CRC校验错误"
        },
        {
            "_input": "68 10 00 00 00 00 01 04 01 00 00 01 00 00 00 00 1A 16",
            "_status": "成功",
            "摘要": "读取数据响应",
            "_table_data": [
                ["  协议控制信息", "68", "帧起始符", "固定值0x68", 0, 0],
                ["  长度域", "10", "16字节", "用户数据长度", 1, 1],
                ["  控制域", "00 00 00 00", "模块→主站", "上行帧", 2, 5],
                ["  地址域", "01 04 01 00", "地址: 0x01040100", "模块地址", 6, 9],
                ["  数据标识", "00 00 00 00", "无数据", "空数据", 10, 13],
                ["  校验码", "1A", "校验和", "算术和校验", 14, 14],
                ["  结束符", "16", "帧结束符", "固定值0x16", 15, 15],
            ]
        }
    ]
    
    exporter = EnhancedBatchResultExporter()
    
    # 验证
    errors = exporter.validate_export_data(sample_results)
    if errors:
        print(f"验证错误: {errors}")
    else:
        print("✓ 数据验证通过")
    
    # JSON 导出
    json_path = exporter.export_to_json(sample_results, "南网协议")
    print(f"✓ JSON 导出: {json_path}")
    
    # Excel 导出（Sheet1 汇总 + Sheet2 详细）
    try:
        excel_path = exporter.export_to_excel(sample_results, "南网协议")
        print(f"✓ Excel 导出: {excel_path}")
        print(f"  - Sheet1: 汇总表（{len(sample_results)} 帧）")
        print(f"  - Sheet2: 详细解析（每帧按字段逐行展开）")
    except ImportError as e:
        print(f"✗ Excel 导出失败: {e}")
    
    # CSV 导出
    try:
        csv_path = exporter.export_to_csv(sample_results, "南网协议")
        print(f"✓ CSV 导出: {csv_path}")
    except ImportError as e:
        print(f"✗ CSV 导出失败: {e}")
    
    print("\n✓ 测试完成")